"""
Take two Excel files and output the differences between the two versions for each panel

"""
import json,requests,sys,datetime,openpyxl


# The user can input the filename of the previous version to be compared at the command line
# panels_prev = sys.argv[1]
panels_prev = r"C:\Users\Verity Fryer\PyCharmProjects\panelapp_queries\outputs\panel_versions_list_20170523.xlsx"

from openpyxl import load_workbook
wb = load_workbook(panels_prev)

sheet = wb['Panels']

for row in range(2, sheet.max_row + 1):
    panel_name_prev = sheet['A' + str(row)].value
    panel_id_prev = sheet['B' + str(row)].value
    panel_version_prev = sheet['C' + str(row)].value
    print(panel_name_prev,panel_id_prev,panel_version_prev)


# open current panels version
# capture current version number for each panel
# for each panel in the current version, find the corresponding panel in the previous version
# if a panel doesn't exist in the previous version, append to the end of the file
# capture date of current version
# capture total_genes_curr

# open previous panels version
# capture previous version number for each panel
# if a panel no longer exists in the previous version, this still needs to be recorded
# capture date of previous version
# capture total_genes_prev

# for a gene within in a panel
#    if gene_name in prev but not curr:
#        retired_genes +1
#    elif gene_name in curr but not prev:
#        new_genes +1
#    elif gene_status_prev == gene_status_curr:
#        unchanged_count +1
#    elif gene_status_prev != gene_status_curr:
#         if gene_status_prev == "HighEvidence" and gene_status_curr == "LowEvidence":
#              green_to_red +1
#         elif gene_status_prev == "LowEvidence" and gene_status_curr == "HighEvidence"
#              red_to_green +1
#         else:
#             pass
#     else:
#          pass